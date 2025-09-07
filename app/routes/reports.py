from collections import defaultdict
import os
import tempfile
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.security import OAuth2PasswordBearer
import openpyxl
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
from app.database.data import supabase
from app.schemas.schemas import ClientReportRequest, ClientReportRequestTimeEntries, ClientTasksBillingRequest, InvoiceByHoursRequest, InvoiceByPercentageRequest, InvoiceFilterRequest, ReportRequest, TaskReportRequest, TaskTimeEntriesRequest, ComprehensiveReportRequest, SimplifiedReportRequest
from app.services.utils import role_required
import pandas as pd
import io
from fastapi.responses import FileResponse, StreamingResponse
from jinja2 import Environment, FileSystemLoader  
from weasyprint import HTML


env = Environment(loader=FileSystemLoader("app/templates"))

router = APIRouter(prefix="/reports", tags=["Reportes"])
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/users/login")


def parse_datetime(date_str):
    """Parse a datetime string with or without milliseconds."""
    try:
        # Try parsing with milliseconds
        return datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S.%f")
    except ValueError:
        # Fallback to parsing without milliseconds
        return datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%S")


def get_billing_type_display(billing_type: str) -> str:
    """Convert database billing type to display name"""
    billing_type_mapping = {
        "fija": "Mensualidad",
        "tarifa_fija": "Tarifa Fija", 
        "hourly": "Por Hora"
    }
    return billing_type_mapping.get(billing_type, billing_type)


def format_hours_to_hhmm(hours: float) -> str:
    """Convert decimal hours to HH:MM format"""
    if hours is None or hours == 0:
        return "00:00"
    
    total_minutes = int(hours * 60)
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{h:02d}:{m:02d}"


def format_currency(amount: float) -> str:
    """Format currency amount as whole number with dot separators"""
    if amount is None:
        return "0"
    return f"{int(round(amount)):,}".replace(",", ".")


@router.post("/hours_by_client/", response_model=List[dict])
async def get_hours_by_client(
    request: ReportRequest,
    user: dict = Depends(role_required(["socio", "senior", "consultor"]))
):
    """get the report of the client between two dates"""
    
    response = supabase.table("time_entries") \
        .select("task_id, duration, start_time, end_time") \
        .gte("start_time", request.start_date.isoformat()) \
        .lte("end_time", request.end_date.isoformat()) \
        .execute()

    if not response.data:
        return []
   
    task_response = supabase.table("tasks").select("id, client_id, title").execute()
    task_dict = {task["id"]: {"client_id": task["client_id"], "title": task["title"]} for task in task_response.data}

    client_hours = {}
   
    for entry in response.data:
        task_id = entry["task_id"]
        task_info = task_dict.get(task_id, None)
        if task_info:
            client_id = task_info["client_id"]
            task_title = task_info["title"]
            
            if client_id not in client_hours:
                client_hours[client_id] = {"total_hours": 0, "tasks": {}}
            
            duration = entry.get("duration", 0) or 0  # Handle None values
            client_hours[client_id]["total_hours"] += duration
            
            if task_id not in client_hours[client_id]["tasks"]:
                client_hours[client_id]["tasks"][task_id] = {"title": task_title, "hours": 0}
            client_hours[client_id]["tasks"][task_id]["hours"] += duration

    client_response = supabase.table("clients").select("id, name").execute()
    client_dict = {client["id"]: client["name"] for client in client_response.data}

    report_data = []
    for client_id, data in client_hours.items():
        client_name = client_dict.get(client_id, "Desconocido")
        total_hours = round(data["total_hours"], 2) 
        
        tasks = [{"task_id": task_id, "Título": task_data["title"], "Horas": round(task_data["hours"], 2)} 
                 for task_id, task_data in data["tasks"].items()]
        
        report_data.append({
            "Cliente": client_name,
            "Total Horas": total_hours,
            "Tareas": tasks
        })

    return report_data

@router.post("/download_report")
async def download_report(
    request: ReportRequest,
    user: dict = Depends(role_required(["socio", "senior", "consultor"]))
):
    """Download the report of clients"""

    # Obtener datos del cliente
    report_data = await get_hours_by_client(request, user)

    # Obtener tareas con las nuevas columnas
    task_response = supabase.table("tasks").select("id, assignment_date, due_date, area").execute()
    task_dict = {task["id"]: task for task in task_response.data}

    # Preparar datos para el archivo Excel
    excel_data = []
    for entry in report_data:
        client_name = entry["Cliente"]
        total_hours = entry["Total Horas"]

        if not entry["Tareas"]:
            excel_data.append({
                "Cliente": client_name,
                "Total Horas": total_hours,
                "Tarea": "",
                "Horas por Tarea": "",
                "Fecha de Asignación": "",
                "Fecha de Vencimiento": "",
                "Área": ""
            })
            continue

        for idx, task in enumerate(entry["Tareas"]):
            # Buscar información de la tarea usando task_id
            task_id = task.get("task_id")
            task_info = task_dict.get(task_id)

            excel_data.append({
                "Cliente": client_name if idx == 0 else "",
                "Total Horas": total_hours if idx == 0 else "",
                "Tarea": task.get("Título", ""),
                "Horas por Tarea": task.get("Horas", ""),
                "Fecha de Asignación": task_info.get("assignment_date", "")[:10] if task_info and task_info.get("assignment_date") else "",
                "Fecha de Vencimiento": task_info.get("due_date", "")[:10] if task_info and task_info.get("due_date") else "",
                "Área": task_info.get("area", "") if task_info else ""
            })

    # Crear archivo Excel
    df = pd.DataFrame(excel_data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Reporte", index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets["Reporte"]

        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'fg_color': '#D7E4BC',
            'border': 1
        })

        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        # Time/duration cell format for hours displayed as [hh]:mm
        time_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '[hh]:mm'
        })

        # Ajustar el ancho de las columnas
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(1, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 20)

        worksheet.set_column('A:A', 30)  # Ajusta el ancho de la columna "Cliente"
        worksheet.set_column('C:C', 30)  # Ajusta el ancho de la columna "Tarea"
        worksheet.set_column('E:E', 25)  # Ajusta el ancho de la columna "Fecha de Asignación"
        worksheet.set_column('F:F', 25)  # Ajusta el ancho de la columna "Fecha de Vencimiento"
        worksheet.set_column('G:G', 20)  # Ajusta el ancho de la columna "Área"

        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        worksheet.merge_range('A1:G1', 'Reporte de Horas por Cliente', title_format)

        for row in range(len(df)):
            for col in range(len(df.columns)):
                value = df.iloc[row, col]
                col_name = df.columns[col]
                # Apply [hh]:mm formatting for duration columns
                if col_name in ["Total Horas", "Horas por Tarea"] and value not in (None, ""):
                    try:
                        hours = float(value)
                        worksheet.write_number(row + 2, col, hours / 24.0, time_format)
                    except Exception:
                        worksheet.write(row + 2, col, value, cell_format)
                else:
                    worksheet.write(row + 2, col, value, cell_format)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_horas_{request.start_date.date()}_{request.end_date.date()}.xlsx"}
    )

@router.post("/download_client_report")
async def download_client_report(
    request: ClientReportRequest,
    user: dict = Depends(role_required(["socio", "senior", "consultor"]))
):
    """Download the report of a specific client"""

    # Obtener datos del cliente
    client_response = supabase.table("clients") \
        .select("id, name") \
        .eq("id", request.client_id) \
        .execute()
    if not client_response.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    client_name = client_response.data[0]["name"]

    # Obtener tareas con las nuevas columnas
    task_response = supabase.table("tasks") \
        .select("id, client_id, title, assignment_date, billing_type ,due_date, area") \
        .eq("client_id", request.client_id) \
        .execute()
    task_dict = {task["id"]: task for task in task_response.data}

    # Preparar datos para el archivo Excel
    excel_data = []
    for task_id, task_info in task_dict.items():
        # Traducir el tipo de facturación a español
        billing_type = task_info['billing_type']
        if billing_type == "hourly":
            billing_type_spanish = "Por Hora"
        elif billing_type == "fija":
            billing_type_spanish = "Mensualidad"
        elif billing_type == "tarifa_fija":
            billing_type_spanish = "Tarifa Fija"
        else:
            billing_type_spanish = billing_type or "No definido"
        
        excel_data.append({
            "Cliente": client_name,
            "Asunto": task_info["title"],
            "Facturacion": billing_type_spanish,
            "Fecha de Asignación": task_info.get("assignment_date", "")[:10] if task_info and task_info.get("assignment_date") else "",
            "Fecha de Vencimiento": task_info.get("due_date", "")[:10] if task_info and task_info.get("due_date") else "",
            "Área": task_info.get("area", "")
        })

    # Crear archivo Excel
    df = pd.DataFrame(excel_data)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Reporte", index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets["Reporte"]

        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'fg_color': '#D7E4BC',
            'border': 1
        })

        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        for col_num, value in enumerate(df.columns.values):
            worksheet.write(1, col_num, value, header_format)
            worksheet.set_column(col_num, col_num, 15)

        worksheet.set_column('A:A', 30)  
        worksheet.set_column('C:C', 30)  

        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        worksheet.merge_range('A1:E1', f'Reporte de Horas para {client_name}', title_format)

        for row in range(len(df)):
            for col in range(len(df.columns)):
                value = df.iloc[row, col]
                worksheet.write(row + 2, col, value, cell_format)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_cliente_{client_name}_{request.start_date.date()}_{request.end_date.date()}.xlsx"}
    )

@router.post("/download_task_report")
async def download_task_report(
    request: TaskReportRequest,
    user: dict = Depends(role_required(["socio", "senior", "consultor"]))
):
    """Download a detailed report of time entries for a specific task (with area from tasks.area)"""

    # Obtener la tarea
    start_date = request.start_date.strftime("%Y-%m-%d")
    end_date = request.end_date.strftime("%Y-%m-%d")

    # Obtener time entries relacionados
    entries_response = supabase.table("time_entries") \
        .select("description, start_time, duration, user_id") \
        .gte("start_time", start_date) \
        .lte("end_time", end_date) \
        .eq("task_id", request.task_id) \
        .execute()

    entries = entries_response.data
    if not entries:
        raise HTTPException(status_code=404, detail="No hay registros de tiempo para esta tarea")

    # Obtener información de la tarea (ahora incluye area)
    task_response = supabase.table("tasks").select("id, client_id, title, area").eq("id", request.task_id).execute()
    if not task_response.data:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    task_data = task_response.data[0]

    # Obtener información del cliente
    client_response = supabase.table("clients").select("id, name").eq("id", task_data["client_id"]).execute()
    if not client_response.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    client_data = client_response.data[0]

    # Obtener información de los usuarios
    user_ids = list(set([entry["user_id"] for entry in entries]))
    users_response = supabase.table("users").select("id, username, role, cost, cost_per_hour_client").in_("id", user_ids).execute()
    if not users_response.data:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    user_dict = {user["id"]: user for user in users_response.data}

    excel_data = []
    for entry in entries:
        user_data = user_dict[entry["user_id"]]

        # Calcular tiempo trabajado
        tiempo_trabajado = float(entry.get("duration", 0) or 0)

        # Calcular tarifa horaria y valor de la hora valorizada
        tarifa_horaria = user_data["cost_per_hour_client"]
        total = tarifa_horaria * tiempo_trabajado

        row = {
            "Abogado": user_data["username"],
            "Cargo": user_data["role"],
            "Cliente": client_data["name"],
            "Asunto": task_data["title"],
            "Trabajo": entry["description"],
            "Área": task_data.get("area", ""),
            "Fecha Trabajo": entry["start_time"][:10],
            "Modo de Facturación": "Por Hora",
            # Keep numeric hours here; we'll apply [hh]:mm number format when writing the cell
            "Tiempo Trabajado": tiempo_trabajado,
            "Tarifa Horaria": format_currency(tarifa_horaria),
            "Moneda": "COP",
            "Total": format_currency(total),
            "Estado de facturación": entry.get("facturado", "no hay datos")
        }

        excel_data.append(row)

    # Crear un nuevo libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entradas"

    # Estilos
    header_fill = openpyxl.styles.PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
    header_font = openpyxl.styles.Font(bold=True)
    header_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    def format_hours(hours):
        """Convert decimal hours to 'HH:MM' format"""
        total_minutes = int(hours * 60)
        h = total_minutes // 60
        m = total_minutes % 60
        return f"{h:02d}:{m:02d}"

    # Escribir el título
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"Desglose de Tiempos - {task_data['title']}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Escribir los encabezados (incluye Área)
    headers = [
        "Abogado", "Cargo", "Cliente", "Asunto", "Trabajo", "Área",
        "Fecha Trabajo", "Modo de Facturación", "Tiempo Trabajado",
        "Tarifa Horaria", "Moneda", "Total", "Facturado"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escribir los datos
    for row_num, row in enumerate(excel_data, 3):
        for col_num, value in enumerate(row.values(), 1):
            # Format "Tiempo Trabajado" column (column I, now 9th column) using Excel [hh]:mm format
            if col_num == 9:  # Column I is the 9th column
                try:
                    hours_val = float(value) if value not in (None, "") else 0.0
                except Exception:
                    # Attempt to parse HH:MM strings if present
                    try:
                        parts = str(value).split(":")
                        hours_val = int(parts[0]) + int(parts[1]) / 60.0
                    except Exception:
                        hours_val = 0.0
                cell = ws.cell(row=row_num, column=col_num, value=hours_val / 24.0)
                cell.number_format = "[hh]:mm"
                cell.alignment = cell_alignment
                cell.border = thin_border
                continue
            # Format "Modo de Facturación" column (column H, now 8th column)
            elif col_num == 8:  # Column H is the 8th column
                value = "Por Hora" if value == "hourly" else value
            # Format numeric columns with commas
            elif col_num in [10, 12]:  # Tarifa Horaria and Total columns
                if isinstance(value, (int, float)):
                    value = f"{value:,.2f}"
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Add a total row at the bottom
    total_sum = sum(entry["Total"] for entry in excel_data)
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "Total:", f"{round(total_sum, 2):,.2f}", ""])

    # Ajustar el ancho de las columnas (incluye Área)
    ws.column_dimensions['A'].width = 30  # Abogado
    ws.column_dimensions['B'].width = 20  # Cargo
    ws.column_dimensions['C'].width = 30  # Cliente
    ws.column_dimensions['D'].width = 30  # Asunto
    ws.column_dimensions['E'].width = 30  # Trabajo
    ws.column_dimensions['F'].width = 15  # Área
    ws.column_dimensions['G'].width = 15  # Fecha Trabajo
    ws.column_dimensions['H'].width = 15  # Modo de Facturación
    ws.column_dimensions['I'].width = 15  # Tiempo Trabajado
    ws.column_dimensions['J'].width = 15  # Tarifa Horaria
    ws.column_dimensions['K'].width = 10  # Moneda
    ws.column_dimensions['L'].width = 15  # Total
    ws.column_dimensions['M'].width = 15  # Facturado

    # Guardar el libro de trabajo en un BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear la respuesta de transmisión
    filename = f"desglose_tarea_{''.join(c if c.isalnum() or c == '_' else '' for c in task_data['title'])}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# @router.post("/download_task_report_old")
# async def download_task_report_old(
#     request: TaskReportRequest,
#     user: dict = Depends(role_required(["socio", "senior", "consultor"]))
# ):
#     """Download a detailed report of time entries for a specific task (with area from tasks.area) - OLD VERSION WITH VALUE_PER_SET_HOURS"""

#     # Obtener la tarea
#     start_date = request.start_date.strftime("%Y-%m-%d")
#     end_date = request.end_date.strftime("%Y-%m-%d")

#     # Obtener time entries relacionados
#     entries_response = supabase.table("time_entries") \
#         .select("description, start_time, duration, user_id") \
#         .gte("start_time", start_date) \
#         .lte("end_time", end_date) \
#         .eq("task_id", request.task_id) \
#         .execute()

#     entries = entries_response.data
#     if not entries:
#         raise HTTPException(status_code=404, detail="No hay registros de tiempo para esta tarea")

#     # Obtener información de la tarea (ahora incluye area)
#     task_response = supabase.table("tasks").select("id, client_id, title, monthly_limit_hours_tasks, area").eq("id", request.task_id).execute()
#     if not task_response.data:
#         raise HTTPException(status_code=404, detail="Tarea no encontrada")
#     task_data = task_response.data[0]

#     # Obtener información del cliente
#     client_response = supabase.table("clients").select("id, name").eq("id", task_data["client_id"]).execute()
#     if not client_response.data:
#         raise HTTPException(status_code=404, detail="Cliente no encontrado")
#     client_data = client_response.data[0]

#     # Obtener información de los usuarios
#     user_ids = list(set([entry["user_id"] for entry in entries]))
#     users_response = supabase.table("users").select("id, username, role, cost, cost_per_hour_client").in_("id", user_ids).execute()
#     if not users_response.data:
#         raise HTTPException(status_code=404, detail="Usuario no encontrado")
#     user_dict = {user["id"]: user for user in users_response.data}

#     # Inicializar variables para el cálculo de tarifas
#     total_set_hours_value = request.value_per_set_hours
#     total_time_entries_duration = 0
    
#     excel_data = []
#     limit_hours_entries = []
#     over_limit_total = 0
#     for entry in entries:
#         user_data = user_dict[entry["user_id"]]

#         # Calcular tiempo trabajado
#         tiempo_trabajado = round(float(entry.get("duration", 0) or 0), 2)

#         # Calcular tarifa horaria y valor de la hora valorizada
#         tarifa_horaria = round(user_data["cost_per_hour_client"], 2)
#         total = round(tarifa_horaria * tiempo_trabajado, 2)
#         cost_to_firm = round(user_data["cost"] * tiempo_trabajado, 2)

#         row = {
#             "Abogado": user_data["username"],
#             "Cargo": user_data["role"],
#             "Cliente": client_data["name"],
#             "Asunto": task_data["title"],
#             "Trabajo": entry["description"],
#             "Área": task_data.get("area", ""),
#             "Fecha Trabajo": entry["start_time"][:10],
#             "Modo de Facturación": "Asesoría Mensual" if total_set_hours_value is not 0 else "Por Hora",
#             "Tiempo Trabajado": tiempo_trabajado,
#             "Tarifa Horaria": tarifa_horaria,
#             "Moneda": "COP",
#             "Total": total,
#             "Facturado": entry.get("facturado", "no hay datos")
#         }

#         if total_set_hours_value is not 0:
#             if total_time_entries_duration < task_data["monthly_limit_hours_tasks"]:
#                 total_time_entries_duration += tiempo_trabajado
#                 limit_hours_entries.append(row)
#             else:
#                 over_limit_total += cost_to_firm
#                 excel_data.append(row)
#         else:
#             excel_data.append(row)

#     # Add the limit hours entries to the excel data
#     if total_set_hours_value is not 0:
#         excel_data = limit_hours_entries + excel_data

#     # Crear un nuevo libro de trabajo
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Entradas"

#     # Estilos
#     header_fill = openpyxl.styles.PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
#     header_font = openpyxl.styles.Font(bold=True)
#     header_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
#     cell_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
#     thin_border = openpyxl.styles.Border(
#         left=openpyxl.styles.Side(style='thin'),
#         right=openpyxl.styles.Side(style='thin'),
#         top=openpyxl.styles.Side(style='thin'),
#         bottom=openpyxl.styles.Side(style='thin')
#     )

#     def format_hours(hours):
#         """Convert decimal hours to 'HH:MM' format"""
#         total_minutes = int(hours * 60)
#         h = total_minutes // 60
#         m = total_minutes % 60
#         return f"{h:02d}:{m:02d}"

#     # Escribir el título
#     ws.merge_cells('A1:M1')
#     title_cell = ws['A1']
#     title_cell.value = f"Desglose de Tiempos - {task_data['title']}"
#     title_cell.font = openpyxl.styles.Font(bold=True, size=14)
#     title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

#     # Escribir los encabezados (incluye Área)
#     headers = [
#         "Abogado", "Cargo", "Cliente", "Asunto", "Trabajo", "Área",
#         "Fecha Trabajo", "Modo de Facturación", "Tiempo Trabajado",
#         "Tarifa Horaria", "Moneda", "Total", "Facturado"
#     ]
#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=2, column=col_num, value=header)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = header_alignment
#         cell.border = thin_border

#     # Escribir los datos
#     for row_num, row in enumerate(excel_data, 3):
#         for col_num, value in enumerate(row.values(), 1):
#             # Format "Tiempo Trabajado" column (column I, now 9th column)
#             if col_num == 9:  # Column I is the 9th column
#                 value = format_hours(float(value))
#             # Format "Modo de Facturación" column (column H, now 8th column)
#             elif col_num == 8:  # Column H is the 8th column
#                 value = "Por Hora" if value == "hourly" else value
#             # Format numeric columns with commas
#             elif col_num in [10, 12]:  # Tarifa Horaria and Total columns
#                 if isinstance(value, (int, float)):
#                     value = f"{value:,.2f}"
#             cell = ws.cell(row=row_num, column=col_num, value=value)
#             cell.alignment = cell_alignment
#             cell.border = thin_border

#     # Add a total row at the bottom
#     total_sum = sum(entry["Total"] for entry in excel_data)
#     if total_set_hours_value > 0:
#         ws.append(["", "", "", "", "", "", "", "", "", "", "", "Total Horas Adicionales:", f"{round(total_sum, 2):,.2f}", ""])
#     else:
#         ws.append(["", "", "", "", "", "", "", "", "", "", "", "Total:", f"{round(total_sum, 2):,.2f}", ""])

#     # Ajustar el ancho de las columnas (incluye Área)
#     ws.column_dimensions['A'].width = 30  # Abogado
#     ws.column_dimensions['B'].width = 20  # Cargo
#     ws.column_dimensions['C'].width = 30  # Cliente
#     ws.column_dimensions['D'].width = 30  # Asunto
#     ws.column_dimensions['E'].width = 30  # Trabajo
#     ws.column_dimensions['F'].width = 15  # Área
#     ws.column_dimensions['G'].width = 15  # Fecha Trabajo
#     ws.column_dimensions['H'].width = 15  # Modo de Facturación
#     ws.column_dimensions['I'].width = 15  # Tiempo Trabajado
#     ws.column_dimensions['J'].width = 15  # Tarifa Horaria
#     ws.column_dimensions['K'].width = 10  # Moneda
#     ws.column_dimensions['L'].width = 15  # Total
#     ws.column_dimensions['M'].width = 15  # Facturado

#     # Guardar el libro de trabajo en un BytesIO
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)

#     # Crear la respuesta de transmisión
#     filename = f"desglose_tarea_{''.join(c if c.isalnum() or c == '_' else '' for c in task_data['title'])}.xlsx"
#     return StreamingResponse(
#         output,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": f"attachment; filename={filename}"}
#     )


@router.post("/get_time_entries")
async def get_time_entries(data: ClientReportRequestTimeEntries, user: dict = Depends(role_required(["socio", "senior", "consultor"]))):
    """ get the time entries by client """

    try:
        start_date = data.start_date.strftime("%Y-%m-%d")
        end_date = data.end_date.strftime("%Y-%m-%d")

        response = (
            supabase.table("time_entries")
            .select("start_time, end_time, task_id, tasks(client_id, clients(name))")
            .gte("start_time", start_date)
            .lte("start_time", end_date)
            .execute()
        )

        if not response.data:
            return []

        time_entries_by_date = {}

        for entry in response.data:
            start_time = parse_datetime(entry["start_time"])
            end_time = parse_datetime(entry["end_time"])
            duration_hours = (end_time - start_time).total_seconds() / 3600

            date_key = start_time.strftime("%Y-%m-%d")
            client_name = entry["tasks"]["clients"]["name"]

            if date_key not in time_entries_by_date:
                time_entries_by_date[date_key] = {}

            if client_name not in time_entries_by_date[date_key]:
                time_entries_by_date[date_key][client_name] = 0

            time_entries_by_date[date_key][client_name] += duration_hours

        result = [
            {"date": date, "client": client, "hours": hours}
            for date, clients in time_entries_by_date.items()
            for client, hours in clients.items()
        ]

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener los registros de tiempo: {str(e)}")

@router.post("/task_time_entries")
async def get_task_time_entries(
    request: TaskTimeEntriesRequest,
    user: dict = Depends(role_required(["socio", "senior"]))
):
    """Get time entries for a specific task filtered by date range and facturado status"""
    
    # Get time entries
    if request.facturado:
        entries_response = supabase.table("time_entries") \
            .select("id,description, start_time, duration, facturado, user_id") \
            .gte("start_time", request.start_date.strftime("%Y-%m-%d")) \
            .lte("end_time", request.end_date.strftime("%Y-%m-%d")) \
            .eq("task_id", request.task_id) \
            .eq("facturado", request.facturado) \
            .order("start_time", desc=False) \
            .execute()
    else:
        entries_response = supabase.table("time_entries") \
            .select("id,description, start_time, duration, facturado, user_id") \
            .gte("start_time", request.start_date.strftime("%Y-%m-%d")) \
            .lte("end_time", request.end_date.strftime("%Y-%m-%d")) \
            .eq("task_id", request.task_id) \
            .order("start_time", desc=False) \
            .execute()
        
    if not entries_response.data:
        return []

    # Get task info
    task_response = supabase.table("tasks").select("id, client_id, title, area").eq("id", request.task_id).execute()
    if not task_response.data:
        return []
    task_data = task_response.data[0]

    # Get client info
    client_response = supabase.table("clients").select("name").eq("id", task_data["client_id"]).execute()
    if not client_response.data:
        return []
    client_name = client_response.data[0]["name"]

    # Process entries
    result = []
    for entry in entries_response.data:
        # Get user info
        user_response = supabase.table("users").select("username, role, cost_per_hour_client").eq("id", entry["user_id"]).execute()
        if not user_response.data:
            continue
        user_data = user_response.data[0]

        # Calculate rates and totals
        tarifa_horaria = user_data["cost_per_hour_client"]
        tiempo_trabajado = round(entry.get("duration", 0) or 0, 2)  # Handle None values
        total = tarifa_horaria * tiempo_trabajado

        # Format time
        horas = int(tiempo_trabajado)
        minutos = int((tiempo_trabajado - horas) * 60)
        tiempo_formateado = f"{horas:02d}:{minutos:02d}"

        result.append({
            "id": entry.get("id"),  
            "abogado": user_data["username"],
            "cargo": user_data["role"],
            "cliente": client_name,
            "trabajo": entry["description"],
            "fecha_trabajo": entry["start_time"][:10],
            "tiempo_trabajado": tiempo_formateado,
            "tarifa_horaria": tarifa_horaria,
            "moneda": "COP",
            "total": total,
            "facturado": entry["facturado"]
        })

    return result


@router.post("/task_time_entries_summary")
async def get_task_time_entries_summary(
    request: TaskTimeEntriesRequest,
    user: dict = Depends(role_required(["socio", "senior"]))
):
    """Get a summary of time entries for a specific task with user rates and totals"""
    
    # Get time entries
    if request.facturado:
        entries_response = supabase.table("time_entries") \
            .select("id, description, start_time, duration, facturado, user_id") \
            .gte("start_time", request.start_date.strftime("%Y-%m-%d")) \
            .lte("end_time", request.end_date.strftime("%Y-%m-%d")) \
            .eq("task_id", request.task_id) \
            .eq("facturado", request.facturado) \
            .order("start_time", desc=False) \
            .execute()
    else:
        entries_response = supabase.table("time_entries") \
            .select("id, description, start_time, duration, facturado, user_id") \
            .gte("start_time", request.start_date.strftime("%Y-%m-%d")) \
            .lte("end_time", request.end_date.strftime("%Y-%m-%d")) \
            .eq("task_id", request.task_id) \
            .order("start_time", desc=False) \
            .execute()
        
    if not entries_response.data:
        return []

    # Get task info including coin
    task_response = supabase.table("tasks").select("id, client_id, title, area, coin").eq("id", request.task_id).execute()
    if not task_response.data:
        return []
    task_data = task_response.data[0]

    # Get client info
    client_response = supabase.table("clients").select("name").eq("id", task_data["client_id"]).execute()
    if not client_response.data:
        return []
    client_name = client_response.data[0]["name"]

    # Get unique user IDs from entries
    user_ids = list(set([entry["user_id"] for entry in entries_response.data]))
    
    # Get user info with rates (only COP rates exist)
    users_response = supabase.table("users").select("id, username, role, cost_per_hour_client").in_("id", user_ids).execute()
    if not users_response.data:
        return []
    
    user_dict = {user["id"]: user for user in users_response.data}

    # Process entries and group by user
    user_summary = {}
    for entry in entries_response.data:
        user_id = entry["user_id"]
        user_data = user_dict.get(user_id)
        
        if not user_data:
            continue
            
        tiempo_trabajado = round(entry.get("duration", 0) or 0, 2)
        
        # Get the COP rate and convert to USD if needed
        tarifa_cop = user_data.get("cost_per_hour_client", 0)
        task_coin = task_data.get("coin", "COP")
        
        if task_coin == "USD":
            # Convert COP to USD using exchange rate from request
            exchange_rate = request.exchange_rate or 4000  # Use request exchange rate or default
            tarifa_horaria = round(tarifa_cop / exchange_rate, 2)
            moneda = "USD"
        else:
            tarifa_horaria = tarifa_cop
            moneda = "COP"
        
        total = round(tarifa_horaria * tiempo_trabajado, 2)
        
        if user_id not in user_summary:
            user_summary[user_id] = {
                "username": user_data["username"],
                "role": user_data["role"],
                "tiempo_trabajado": 0,
                "tarifa_horaria": tarifa_horaria,
                "moneda": moneda,
                "total": 0
            }
        
        user_summary[user_id]["tiempo_trabajado"] += tiempo_trabajado
        user_summary[user_id]["total"] += total

    # Convert to list and round totals
    result = []
    for user_data in user_summary.values():
        result.append({
            "username": user_data["username"],
            "role": user_data["role"],
            "tiempo_trabajado": round(user_data["tiempo_trabajado"], 2),
            "tarifa_horaria": round(user_data["tarifa_horaria"], 2),
            "moneda": user_data["moneda"],
            "total": round(user_data["total"], 2)
        })

    return result

@router.post("/client_tasks_billing")
async def get_client_tasks_billing(
    request: ClientTasksBillingRequest,
    user: dict = Depends(role_required(["socio", "senior"]))
):
    """Get all tasks for a client with billing details, package hours calculations, and additional charges"""
    
    # Get all tasks for the client
    tasks_response = supabase.table("tasks") \
        .select("id, title, billing_type, area, coin, asesoria_tarif, total_value, permanent, monthly_limit_hours_tasks") \
        .eq("client_id", request.client_id) \
        .execute()
    
    if not tasks_response.data:
        return []
    
    tasks = tasks_response.data
    result = []
    
    for task in tasks:
        task_id = task["id"]
        
        # Get time entries for this task within the specified date range
        time_entries_response = supabase.table("time_entries") \
            .select("duration, user_id, start_time") \
            .eq("task_id", task_id) \
            .gte("start_time", request.start_date.isoformat()) \
            .lte("start_time", request.end_date.isoformat()) \
            .execute()
        
        total_time = 0
        total_generated = 0
        
        if time_entries_response.data:
            # Get user rates for this task
            user_ids = list(set([entry["user_id"] for entry in time_entries_response.data]))
            users_response = supabase.table("users").select("id, cost_per_hour_client").in_("id", user_ids).execute()
            
            if users_response.data:
                user_dict = {user["id"]: user for user in users_response.data}
                
                # Calculate totals
                for entry in time_entries_response.data:
                    duration = entry.get("duration", 0) or 0
                    total_time += duration
                    
                    user_rate = user_dict.get(entry["user_id"], {}).get("cost_per_hour_client", 0)
                    
                    # Convert to USD if task currency is USD (using default exchange rate)
                    if task.get("coin") == "USD":
                        exchange_rate = 4000  # Default exchange rate
                        user_rate = user_rate / exchange_rate
                    
                    total_generated += duration * user_rate
        
        # Round totals
        total_time = round(total_time, 2)
        total_generated = round(total_generated, 2)
        
        # Prepare task data
        task_data = {
            "id": task_id,
            "title": task["title"],
            "billing_type": task["billing_type"],
            "area": task.get("area"),
            "coin": task.get("coin", "COP"),
            "total_time": total_time,
            "total_generated": total_generated,
            "permanent": task.get("permanent", False)
        }
        
        # Add billing-specific fields based on billing type
        if task["billing_type"] == "tarifa_fija":
            # Fixed rate billing
            task_data["asesoria_tarif"] = task.get("asesoria_tarif", 0)
            task_data["total_value"] = task.get("total_value", 0)
            task_data["billing_type_display"] = "Tarifa Fija"
        elif task["billing_type"] == "fija":
            # Monthly subscription billing (mensualidad)
            task_data["asesoria_tarif"] = task.get("asesoria_tarif", 0)
            task_data["monthly_limit_hours"] = task.get("monthly_limit_hours_tasks", 0)
            task_data["billing_type_display"] = "Mensualidad"
            
            # Check if monthly limit hours have been reached
            monthly_limit = task.get("monthly_limit_hours_tasks", 0) or 0
            if monthly_limit > 0:
                if total_time >= monthly_limit:
                    task_data["monthly_limit_reached"] = True
                    task_data["additional_hours"] = total_time - monthly_limit
                    # Calculate additional charges for hours beyond the limit
                    if time_entries_response.data and users_response.data:
                        # Find entries that exceed the monthly limit
                        remaining_entries = []
                        current_time = 0
                        
                        for entry in sorted(time_entries_response.data, key=lambda x: x.get("start_time", "")):
                            duration = entry.get("duration", 0) or 0
                            
                            if current_time + duration <= monthly_limit:
                                # This entry fits completely within the monthly limit
                                current_time += duration
                            else:
                                # This entry exceeds the monthly limit
                                if current_time < monthly_limit:
                                    # Partial entry - only charge the excess portion
                                    excess_duration = (current_time + duration) - monthly_limit
                                    remaining_entries.append({
                                        "duration": excess_duration,
                                        "user_id": entry["user_id"]
                                    })
                                    current_time = monthly_limit  # Mark that we've reached the limit
                                else:
                                    # Full entry is beyond the limit - charge for all hours
                                    remaining_entries.append(entry)
                                    current_time += duration
                        
                        # Debug: Show what we found
                        total_excess_hours = sum(entry['duration'] for entry in remaining_entries)
                        print(f"Debug: Monthly limit: {monthly_limit}h")
                        print(f"Debug: Total time worked: {total_time}h")
                        print(f"Debug: Hours within limit: {monthly_limit}h")
                        print(f"Debug: Hours beyond limit: {total_excess_hours}h")
                        print(f"Debug: Verification: {monthly_limit + total_excess_hours} should equal {total_time}")
                        
                        # Show each entry that exceeds the limit
                        for i, entry in enumerate(remaining_entries):
                            print(f"Debug: Entry {i+1}: User {entry['user_id']} - {entry['duration']}h")
                        
                        # Prepare to calculate additional charges from breakdown
                        # (The actual calculation will be done in the breakdown loop below)
                    
                    # Add detailed breakdown of additional charges
                    task_data["additional_charges_breakdown"] = []
                    breakdown_total = 0  # Track the sum of subtotals
                    
                    for entry in remaining_entries:
                        user_id = entry["user_id"]
                        duration = entry["duration"]
                        
                        # Get the user's rate from the users table
                        user_rate = user_dict.get(user_id, {}).get("cost_per_hour_client", 0)
                        
                        # Validate that we got a valid rate
                        if user_rate is None or user_rate == 0:
                            print(f"Warning: User {user_id} has no rate or rate is 0")
                        
                        # Apply currency conversion if needed
                        if task.get("coin") == "USD":
                            exchange_rate = 4000  # Default COP to USD rate
                            user_rate = user_rate / exchange_rate
                        
                        # Calculate subtotal: time_worked_by_user × user_tariff
                        subtotal = duration * user_rate
                        breakdown_total += subtotal
                        
                        # Debug logging
                        print(f"User {user_id}: {duration}h × ${user_rate}/h = ${subtotal}")
                        
                        # Get user name for display
                        user_name = "Unknown User"
                        if users_response.data:
                            for user in users_response.data:
                                if user["id"] == user_id:
                                    user_name = user.get("username", "Unknown User")
                                    break
                        
                        breakdown_entry = {
                            "user_id": user_id,
                            "user_name": user_name,
                            "hours": duration,
                            "rate_per_hour": user_rate,
                            "subtotal": subtotal
                        }
                        task_data["additional_charges_breakdown"].append(breakdown_entry)
                    
                    # Set additional charge as the sum of all user subtotals
                    # Formula: Σ (time_worked_by_user × user_tariff)
                    task_data["additional_charge"] = round(breakdown_total, 2)
                    task_data["total_with_additional"] = round(task_data["asesoria_tarif"] + breakdown_total, 2)
                    
                    # Validate that our calculation matches the expected excess hours
                    calculated_excess_hours = sum(entry['duration'] for entry in remaining_entries)
                    if abs(calculated_excess_hours - task_data['additional_hours']) > 0.01:
                        print(f"Warning: Calculated excess hours ({calculated_excess_hours}) don't match expected ({task_data['additional_hours']})")
                    
                    # Debug summary
                    print(f"Task {task_id}: Monthly limit {monthly_limit}h, worked {total_time}h")
                    print(f"Expected excess hours: {task_data['additional_hours']}h")
                    print(f"Calculated excess hours: {calculated_excess_hours}h")
                    print(f"Additional charge breakdown: {len(task_data['additional_charges_breakdown'])} entries")
                    print(f"Total additional charge: ${breakdown_total}")
                    print(f"Final total: ${task_data['total_with_additional']}")
                else:
                    task_data["monthly_limit_reached"] = False
                    task_data["additional_hours"] = 0
                    task_data["additional_charge"] = 0
                    task_data["total_with_additional"] = task_data["asesoria_tarif"]
            else:
                # No monthly limit set
                task_data["monthly_limit_reached"] = False
                task_data["additional_hours"] = 0
                task_data["additional_charge"] = 0
                task_data["total_with_additional"] = task_data["asesoria_tarif"]
        elif task["billing_type"] == "hourly":
            # Hourly billing
            task_data["billing_type_display"] = "Por Hora"
            # For hourly tasks, the total is based on time worked
            task_data["total_value"] = total_generated
        else:
            # Unknown billing type
            task_data["billing_type_display"] = task["billing_type"] or "No definido"
        
        # Set final total based on billing type
        if task["billing_type"] == "fija":
            # For mensualidad tasks, use the total with additional charges if applicable
            if "total_with_additional" in task_data:
                task_data["final_total"] = task_data["total_with_additional"]
            else:
                task_data["final_total"] = task_data.get("asesoria_tarif", 0)
        elif task["billing_type"] == "tarifa_fija":
            # For fixed rate tasks, use the fixed rate
            task_data["final_total"] = task_data.get("asesoria_tarif", 0)
        else:
            # For hourly tasks, use the generated total
            task_data["final_total"] = total_generated
        
        result.append(task_data)
    
    return result








@router.post("/invoices/by-hours")
def generate_invoice_by_hours(req: InvoiceByHoursRequest):
    try:
        today = datetime.now()

        if req.currency == "USD" and not req.exchange_rate:
            raise HTTPException(status_code=400, detail="Debe enviar exchange_rate si la moneda es USD")

        # Verificar tarea
        task_resp = supabase.table("tasks").select("id, title, area").eq("id", req.task_id).single().execute()
        if not task_resp or not task_resp.data:
            raise HTTPException(status_code=404, detail="Tarea no encontrada")
        task = task_resp.data

        # Obtener cliente
        client_resp = supabase.table("clients").select("*").eq("id", req.client_id).single().execute()
        if not client_resp or not client_resp.data:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        client = client_resp.data

        # Obtener time entries no facturados para la tarea
        time_resp = supabase.table("time_entries")\
            .select("duration, user_id")\
            .eq("task_id", req.task_id)\
            .eq("facturado", "no")\
            .execute()

        if not time_resp or not time_resp.data:
            raise HTTPException(status_code=400, detail="No hay registros de tiempo sin facturar para esta tarea")
        entries = time_resp.data

        # Obtener usuarios relacionados
        user_ids = list(set([e["user_id"] for e in entries]))
        users_resp = supabase.table("users").select("id, username, cost, cost_per_hour_client").in_("id", user_ids).execute()
        if not users_resp or not users_resp.data:
            raise HTTPException(status_code=400, detail="No se encontraron usuarios relacionados")
        user_map = {u["id"]: u for u in users_resp.data}

        # Calcular totales y detalles para PDF
        subtotal = 0
        total_hours = 0
        tasks_details = []
        for e in entries:
            uid = e["user_id"]
            duration = round(e["duration"], 2)
            rate = user_map[uid]["cost_per_hour_client"]
            total = duration * rate

            if req.currency == "USD":
                rate = round(rate / req.exchange_rate, 2)
                total = round(total / req.exchange_rate, 2)

            subtotal += total
            total_hours += duration

            tasks_details.append({
                "username": user_map[uid]["username"],
                "area": task.get("area", "Sin área"),
                "description": task["title"],
                "duration": duration,
                "rate": rate,
                "total": total
            })

        subtotal = round(subtotal, 2)
        tax = round(subtotal * 0.19, 2) if req.include_tax else 0
        total = round(subtotal + tax, 2)

        # Registrar orden de servicio
        invoice_data = {
            "client_id": req.client_id,
            "start_date": str(today.replace(day=1).date()),
            "end_date": str(today.date()),
            "issued_at": today.isoformat(),
            "billing_type": "hourly",
            "total_hours": total_hours,
            "subtotal": subtotal,
            "tax": tax,
            "total": total,
            "currency": req.currency,
            "exchange_rate": req.exchange_rate,
            "include_tax": req.include_tax
        }
        supabase.table("invoices").insert(invoice_data).execute()

        # Actualizar total facturado en la tarea
        new_total = round(task.get("total_billed", 0) + total, 2)
        update_fields = {
            "total_billed": new_total
        }
        supabase.table("tasks").update(update_fields).eq("id", req.task_id).execute()

        # Marcar time_entries como facturados
        supabase.table("time_entries").update({"facturado": "si"}).eq("task_id", req.task_id).eq("facturado", "no").execute()

        # Renderizar HTML para el PDF
        template = env.get_template("invoice_template.html")
        html_out = template.render(
            client=client,
            date=today.strftime("%Y-%m-%d"),
            start_date=today.replace(day=1).date(),
            end_date=today.date(),
            tasks_details=tasks_details,
            subtotal=subtotal,
            tax=tax,
            total=total,
            currency_symbol="$" if req.currency == "COP" else "USD",
            billing_type="hourly"
        )

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmpfile:
            pdf_path = tmpfile.name
        HTML(string=html_out, base_url=os.getcwd()).write_pdf(pdf_path)

        return FileResponse(
            path=pdf_path,
            filename=f"orden_servicio_{client['name'].replace(' ', '_')}_{today.strftime('%Y-%m-%d')}.pdf",
            media_type="application/pdf"
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")
    

@router.post("/invoices/by-percentage")
def generate_invoice_by_percentage(req: InvoiceByPercentageRequest):
    try:
        if req.currency == "USD" and not req.exchange_rate:
            raise HTTPException(status_code=400, detail="Debe enviar exchange_rate si la moneda es USD")

        task_resp = supabase.table("tasks").select("*", count="exact").eq("id", req.task_id).single().execute()
        if not task_resp or not task_resp.data:
            raise HTTPException(status_code=404, detail="Tarea no encontrada")
        task = task_resp.data

        total_case_value = task.get("total_value")
        if total_case_value is None:
            raise HTTPException(status_code=400, detail="La tarea no tiene un valor total definido")

        billed_raw = total_case_value * (req.percentage / 100)
        subtotal = round(billed_raw if req.currency == "COP" else billed_raw / req.exchange_rate, 2)
        tax = round(subtotal * 0.19, 2)
        total = round(subtotal + tax, 2)

        now = datetime.now()
        currency_symbol = "$" if req.currency == "COP" else "USD"

        # Registrar orden de servicio (invoice)
        invoice_data = {
            "client_id": req.client_id,
            "start_date": str(now.replace(day=1).date()),
            "end_date": str(now.date()),
            "issued_at": now.isoformat(),
            "billing_type": "percentage",
            "percentage": req.percentage,
            "subtotal": subtotal,
            "tax": tax,
            "total": total,
            "payment_type": req.payment_type,
            "total_case_value": total_case_value,
            "currency": req.currency,
            "exchange_rate": req.exchange_rate
        }
        supabase.table("invoices").insert(invoice_data).execute()

        # Actualizar progreso de facturación en la tarea
        new_percentage = round(task.get("percentage_billed", 0) + req.percentage, 2)
        new_total = round(task.get("total_billed", 0) + total, 2)

        update_fields = {
            "percentage_billed": new_percentage,
            "total_billed": new_total
        }

        if new_percentage >= 100:
            supabase.table("time_entries").update({"facturado": "si"}).eq("task_id", req.task_id).execute()
        elif new_percentage > 0:
            supabase.table("time_entries").update({"facturado": "parcialmente"}).eq("task_id", req.task_id).execute()

        supabase.table("tasks").update(update_fields).eq("id", req.task_id).execute()

        # Obtener los datos del cliente para el PDF
        client_resp = supabase.table("clients").select("name").eq("id", req.client_id).single().execute()
        client_name = client_resp.data["name"]

        # Obtener los detalles de la tarea
        task_details = {
            "username": "—",
            "area": task["area"],
            "description": task["title"],
            "duration": "—",
            "rate": "—",
            "total": total
        }

        # Preparar el HTML para el PDF
        # Calcular el valor restante considerando la moneda
        if req.currency == "USD":
            restante = (total_case_value - new_total) / req.exchange_rate
        else:
            restante = total_case_value - new_total
            
        template = env.get_template("invoice_template.html")
        html_out = template.render(
            client=client_resp.data,
            date=now.strftime("%Y-%m-%d"),
            start_date=now.replace(day=1).date(),
            end_date=now.date(),
            tasks_details=[task_details],
            subtotal=subtotal,
            tax=tax,
            total=total,
            currency_symbol=currency_symbol,
            billing_type="percentage",  # Solo porcentaje aquí
            total_facturado=new_total,
            restante=restante
        )

        # Crear el PDF
        pdf_path = "/tmp/invoice.pdf"
        HTML(string=html_out).write_pdf(pdf_path)

        return FileResponse(pdf_path, media_type="application/pdf", filename=f"invoice_{client_name}_{now.strftime('%Y-%m-%d')}.pdf")

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")

@router.get("/invoices/registry")
def get_invoice_registry(
    start_date: str = Query(..., description="Fecha inicial en formato YYYY-MM-DD"),
    end_date: str = Query(..., description="Fecha final en formato YYYY-MM-DD")
):
    try:
        # Obtener facturas en el rango
        response = supabase.table("invoices")\
            .select("*, clients(name)")\
            .gte("issued_at", start_date)\
            .lte("issued_at", end_date)\
            .order("issued_at", desc=True)\
            .execute()

        raw_data = response.data or []

        # Transformar resultados
        facturas = []
        for f in raw_data:
            facturas.append({
                "id": f["id"],
                "issued_at": f["issued_at"],
                "billing_type": f["billing_type"],
                "subtotal": f["subtotal"],
                "tax": f["tax"],
                "total": f["total"],
                "percentage": f.get("percentage"),
                "payment_type": f.get("payment_type"),
                "total_hours": f.get("total_hours"),
                "total_case_value": f.get("total_case_value"),
                "client_name": f["clients"]["name"]
            })

        return facturas

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")

@router.get("/invoices/registry/excel")
async def get_invoice_registry_excel(
    start_date: str = Query(..., description="Fecha inicial en formato YYYY-MM-DD"),
    end_date: str = Query(..., description="Fecha final en formato YYYY-MM-DD")
):
    try:
        # Obtener facturas en el rango
        response = supabase.table("invoices")\
            .select("issued_at, billing_type, subtotal, currency")\
            .gte("issued_at", start_date)\
            .lte("issued_at", end_date)\
            .order("issued_at", desc=True)\
            .execute()

        if not response.data:
            raise HTTPException(status_code=404, detail="No hay facturas en el rango de fechas seleccionado")

        # Preparar datos para Excel
        excel_data = []
        for invoice in response.data:
            # Traducir el tipo de facturación a español
            if invoice["billing_type"] == "hourly":
                billing_type = "Por Hora"
            elif invoice["billing_type"] == "fija":
                billing_type = "Mensualidad"
            elif invoice["billing_type"] == "tarifa_fija":
                billing_type = "Tarifa Fija"
            else:
                billing_type = invoice["billing_type"] or "No definido"
            currency = "COP" if invoice["currency"] == "COP" else "USD"
            
            excel_data.append({
                "Fecha de Emisión": invoice["issued_at"].split("T")[0],
                "Tipo de Facturación": billing_type,
                "Subtotal": invoice["subtotal"],
                "Moneda": currency
            })

        # Crear archivo Excel
        df = pd.DataFrame(excel_data)
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Registro de Facturas", index=False, startrow=1)
            workbook = writer.book
            worksheet = writer.sheets["Registro de Facturas"]

            # Formato para encabezados
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'vcenter',
                'align': 'center',
                'fg_color': '#D7E4BC',
                'border': 1
            })

            # Formato para celdas
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            # Aplicar formatos y ajustar columnas
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(1, col_num, value, header_format)
                worksheet.set_column(col_num, col_num, 20)

            # Ajustar el ancho de las columnas específicas
            worksheet.set_column('A:A', 15)  # Fecha de Emisión
            worksheet.set_column('B:B', 20)  # Tipo de Facturación
            worksheet.set_column('C:C', 15)  # Subtotal
            worksheet.set_column('D:D', 10)  # Moneda

            # Título del reporte
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 14,
                'align': 'center',
                'valign': 'vcenter'
            })
            worksheet.merge_range('A1:D1', 'Registro de Facturas', title_format)

            # Aplicar formato a las celdas de datos
            for row in range(len(df)):
                for col in range(len(df.columns)):
                    value = df.iloc[row, col]
                    worksheet.write(row + 2, col, value, cell_format)

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=registro_facturas_{start_date}_{end_date}.xlsx"}
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")


@router.post("/comprehensive_report")
async def generate_comprehensive_report(
    request: ComprehensiveReportRequest,
    user: dict = Depends(role_required(["socio", "senior", "consultor"]))
):
    """
    Generate comprehensive reports based on type:
    - general: Complete time tracking report with lawyer rates (includes all billing types)
    - tarifa_fija: Fixed rate clients report
    - mensualidad: Monthly subscription clients report
    - hourly: Hourly billing tasks report
    
    If task_id is provided, returns a task-specific report in the format corresponding to the task's billing type:
    - hourly tasks: Hourly billing format with time tracking and rates
    - tarifa_fija tasks: Fixed rate format showing fixed amount
    - fija (mensualidad) tasks: Monthly subscription format with limits and additional charges
    - other billing types: General format as fallback
    """
    try:
        from datetime import datetime, date
        import calendar
        
        # Set default dates to current month if not provided
        if not request.start_date or not request.end_date:
            today = date.today()
            first_day = date(today.year, today.month, 1)
            last_day = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
            start_date = datetime.combine(first_day, datetime.min.time())
            end_date = datetime.combine(last_day, datetime.max.time())
        else:
            start_date = request.start_date
            end_date = request.end_date

        # If task_id is provided, generate task-specific report based on billing type
        if request.task_id:
            return await _generate_task_specific_report(start_date, end_date, request.task_id, request.report_type)
        
        # Otherwise, generate comprehensive reports as before
        if request.report_type == "general":
            return await _generate_general_report(start_date, end_date, request.client_id)
        elif request.report_type == "tarifa_fija":
            return await _generate_fixed_rate_report(start_date, end_date, request.client_id)
        elif request.report_type == "mensualidad":
            return await _generate_monthly_report(start_date, end_date, request.client_id)
        elif request.report_type == "hourly":
            return await _generate_hourly_report(start_date, end_date, request.client_id)
        else:
            raise HTTPException(status_code=400, detail="Tipo de reporte no válido")

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")


async def _generate_general_report(start_date: datetime, end_date: datetime, client_id: Optional[int] = None):
    """Generate the general report with time tracking and lawyer rates"""
    
    # 1) Obtener time entries del rango (todas las modalidades)
    entries_query = supabase.table("time_entries").select("""
        id, duration, start_time, end_time, description, user_id, facturado,
        tasks!inner(id, title, client_id, area, billing_type, facturado, assigned_user_name)
    """).gte("start_time", start_date.isoformat()).lte("end_time", end_date.isoformat())
    if client_id:
        entries_query = entries_query.eq("tasks.client_id", client_id)
    entries_response = entries_query.execute()

    # 2) Obtener todas las tareas del cliente (o de todos los clientes) para incluir las que no tengan tiempos
    tasks_query = supabase.table("tasks").select("id, title, client_id, area, billing_type, assigned_user_name")
    if client_id:
        tasks_query = tasks_query.eq("client_id", client_id)
    tasks_response = tasks_query.execute()

    if not tasks_response.data:
        raise HTTPException(status_code=404, detail="No se encontraron tareas para el filtro seleccionado")

    # 3) Mapear clientes y usuarios necesarios
    entry_client_ids = list(set([e.get("tasks", {}).get("client_id") for e in (entries_response.data or []) if e.get("tasks", {}).get("client_id")]))
    task_client_ids = list(set([t.get("client_id") for t in tasks_response.data if t.get("client_id")]))
    all_client_ids = list(set(entry_client_ids + task_client_ids))

    if all_client_ids:
        clients_response = supabase.table("clients").select("id, name").in_("id", all_client_ids).execute()
        client_dict = {client["id"]: client["name"] for client in (clients_response.data or [])}
    else:
        client_dict = {}

    entry_user_ids = list(set([e.get("user_id") for e in (entries_response.data or []) if e.get("user_id")]))
    users_response = supabase.table("users").select("id, username, role, cost_per_hour_client").in_("id", entry_user_ids).execute() if entry_user_ids else type("obj", (), {"data": []})
    user_dict = {user["id"]: {"username": user["username"], "role": user.get("role", ""), "rate": user.get("cost_per_hour_client", 0)} for user in (users_response.data or [])}

    # 4) Construir filas a partir de time entries existentes
    excel_data = []
    task_ids_with_entries = set()
    for entry in (entries_response.data or []):
        task = entry.get("tasks", {})
        user_id = entry.get("user_id")
        user_info = user_dict.get(user_id, {}) if user_id else {}
        c_id = task.get("client_id")
        client_name = client_dict.get(c_id, "")
        rate = user_info.get("rate", 0) or 0
        duration = entry.get("duration", 0) or 0
        total = rate * duration
        task_ids_with_entries.add(task.get("id"))

        excel_data.append({
            "Abogado": user_info.get("username", ""),
            "Rol": user_info.get("role", ""),
            "Nombre del Cliente": client_name,
            "Asunto": task.get("title", ""),
            "Descripción": entry.get("description", ""),
            "Área": task.get("area", ""),
            "Tipo de facturación": get_billing_type_display(task.get("billing_type", "")),
            "Tiempo reportado": format_hours_to_hhmm(duration),
            "Fecha de reporte": entry.get("start_time", "")[:10] if entry.get("start_time") else "",
            "Tarifa del abogado": format_currency(rate),
            "Total Tarifa x Tiempo": format_currency(total),
            "Estado de facturación": entry.get("facturado", ""),
            "Abogado asignado": task.get("assigned_user_name", "") or "Sin abogado asignado"
        })

    # 5) Agregar filas de tareas sin tiempos
    for t in tasks_response.data:
        if t["id"] in task_ids_with_entries:
            continue
        client_name = client_dict.get(t.get("client_id"), "")
        excel_data.append({
            "Abogado": "",
            "Rol": "",
            "Nombre del Cliente": client_name,
            "Asunto": t.get("title", ""),
            "Descripción": "",
            "Área": t.get("area", ""),
            "Tipo de facturación": get_billing_type_display(t.get("billing_type", "")),
            "Tiempo reportado": format_hours_to_hhmm(0),
            "Fecha de reporte": "",
            "Tarifa del abogado": format_currency(0),
            "Total Tarifa x Tiempo": format_currency(0),
            "Estado de facturación": "",
            "Abogado asignado": t.get("assigned_user_name", "") or "Sin abogado asignado"
        })

    return _create_comprehensive_excel_file(excel_data, "Reporte General", start_date, end_date)


async def _generate_fixed_rate_report(start_date: datetime, end_date: datetime, client_id: Optional[int] = None):
    """Generate the fixed rate clients report"""
    
    # Build query for fixed rate tasks - show all tasks with this billing type
    query = supabase.table("tasks").select("""
        id, title, created_at, billing_type, total_value, facturado, note, assigned_user_name, client_id
    """).eq("billing_type", "tarifa_fija")
    
    # Only filter by client if client_id is provided
    if client_id:
        query = query.eq("client_id", client_id)
    
    response = query.execute()
    
    if not response.data:
        raise HTTPException(status_code=404, detail="No se encontraron tareas con tarifa fija")
    
    # Get client information
    client_ids = list(set([task.get("client_id") for task in response.data if task.get("client_id")]))
    clients_response = supabase.table("clients").select("id, name").in_("id", client_ids).execute()
    client_dict = {client["id"]: client["name"] for client in clients_response.data} if clients_response.data else {}
    
    # Prepare Excel data
    excel_data = []
    for task in response.data:
        client_name = client_dict.get(task.get("client_id"), "")
        
        excel_data.append({
            "Cliente": client_name,
            "Asunto": task.get("title", ""),
            "Fecha de creación": task.get("created_at", "")[:10] if task.get("created_at") else "",
            "Tipo de facturación": get_billing_type_display(task.get("billing_type", "")),
            "Tarifa fija": format_currency(task.get('total_value', 0) or 0),
            "Estado de facturación": task.get("facturado", ""),
            "Nota": task.get("note", ""),
            "Abogado asignado": task.get("assigned_user_name", "") or "Sin abogado asignado"
        })
    
    return _create_comprehensive_excel_file(excel_data, "Reporte Tarifa Fija", start_date, end_date)


async def _generate_monthly_report(start_date: datetime, end_date: datetime, client_id: Optional[int] = None):
    """Generate the monthly subscription clients report"""
    
    # Build query for monthly subscription tasks - show all tasks with this billing type
    query = supabase.table("tasks").select("""
        id, title, area, billing_type, monthly_limit_hours_tasks, asesoria_tarif, facturado, client_id, assigned_user_name
    """).eq("billing_type", "fija")
    
    # Only filter by client if client_id is provided
    if client_id:
        query = query.eq("client_id", client_id)
    
    response = query.execute()
    
    if not response.data:
        raise HTTPException(status_code=404, detail="No se encontraron tareas con mensualidad (billing_type: fija)")
    
    # Get client information
    client_ids = list(set([task.get("client_id") for task in response.data if task.get("client_id")]))
    clients_response = supabase.table("clients").select("id, name").in_("id", client_ids).execute()
    client_dict = {client["id"]: client["name"] for client in clients_response.data} if clients_response.data else {}
    
    # Get time entries for these tasks in the date range with user information
    task_ids = [task["id"] for task in response.data]
    time_response = supabase.table("time_entries").select("""
        task_id, duration, user_id
    """).in_("task_id", task_ids).gte("start_time", start_date.isoformat()).lte("end_time", end_date.isoformat()).execute()
    
    # Get user rates for all users who worked on these tasks
    user_ids = list(set([entry.get("user_id") for entry in time_response.data if entry.get("user_id")]))
    users_response = supabase.table("users").select("id, cost_per_hour_client").in_("id", user_ids).execute()
    user_dict = {user["id"]: user.get("cost_per_hour_client", 0) for user in users_response.data} if users_response.data else {}
    
    # Calculate total time and worked value per task
    task_time = {}
    task_worked_value = {}
    for entry in time_response.data:
        task_id = entry["task_id"]
        duration = entry.get("duration", 0) or 0
        user_id = entry.get("user_id")
        user_rate = user_dict.get(user_id, 0) or 0
        
        if task_id not in task_time:
            task_time[task_id] = 0
            task_worked_value[task_id] = 0
        
        task_time[task_id] += duration
        task_worked_value[task_id] += duration * user_rate
    
    # Prepare Excel data
    excel_data = []
    for task in response.data:
        client_name = client_dict.get(task.get("client_id"), "")
        
        # Calculate values
        total_time = task_time.get(task["id"], 0) or 0
        value_worked = task_worked_value.get(task["id"], 0) or 0
        monthly_rate = task.get("asesoria_tarif", 0) or 0
        difference = monthly_rate - value_worked
        
        excel_data.append({
            "Nombre del Cliente": client_name,
            "Asunto": task.get("title", ""),
            "Área": task.get("area", ""),
            "Tipo de facturación": "Mensualidad",
            "Tiempo reportado": format_hours_to_hhmm(total_time),
            "Valor trabajado": format_currency(value_worked),
            "Límite de horas mensuales": format_hours_to_hhmm(task.get("monthly_limit_hours_tasks", 0) or 0),
            "Tarifa mensualidad": format_currency(monthly_rate),
            "Diferencia": format_currency(difference),
            "Abogado asignado": task.get("assigned_user_name", "") or "Sin abogado asignado"
        })
    
    return _create_comprehensive_excel_file(excel_data, "Reporte Mensualidad", start_date, end_date, 
                                         conditional_formatting=True, difference_column="H")


async def _generate_hourly_report(start_date: datetime, end_date: datetime, client_id: Optional[int] = None):
    """Generate the hourly billing tasks report"""
    
    # 1) Obtener time entries de tareas hourly
    entries_query = supabase.table("time_entries").select("""
        id, duration, start_time, end_time, description, user_id, facturado,
        tasks!inner(id, title, client_id, area, billing_type, facturado, assigned_user_name)
    """).gte("start_time", start_date.isoformat()).lte("end_time", end_date.isoformat()).eq("tasks.billing_type", "hourly")
    if client_id:
        entries_query = entries_query.eq("tasks.client_id", client_id)
    entries_response = entries_query.execute()

    # 2) Obtener todas las tareas hourly (para incluir las que no tengan tiempos)
    tasks_query = supabase.table("tasks").select("id, title, client_id, area, billing_type, assigned_user_name").eq("billing_type", "hourly")
    if client_id:
        tasks_query = tasks_query.eq("client_id", client_id)
    tasks_response = tasks_query.execute()

    if not tasks_response.data:
        raise HTTPException(status_code=404, detail="No se encontraron tareas con facturación por hora")

    # 3) Mapear clientes y usuarios
    entry_client_ids = list(set([e.get("tasks", {}).get("client_id") for e in (entries_response.data or []) if e.get("tasks", {}).get("client_id")]))
    task_client_ids = list(set([t.get("client_id") for t in tasks_response.data if t.get("client_id")]))
    all_client_ids = list(set(entry_client_ids + task_client_ids))

    if all_client_ids:
        clients_response = supabase.table("clients").select("id, name").in_("id", all_client_ids).execute()
        client_dict = {client["id"]: client["name"] for client in (clients_response.data or [])}
    else:
        client_dict = {}

    entry_user_ids = list(set([e.get("user_id") for e in (entries_response.data or []) if e.get("user_id")]))
    users_response = supabase.table("users").select("id, username, role, cost_per_hour_client").in_("id", entry_user_ids).execute() if entry_user_ids else type("obj", (), {"data": []})
    user_dict = {user["id"]: {"username": user["username"], "role": user.get("role", ""), "rate": user.get("cost_per_hour_client", 0)} for user in (users_response.data or [])}

    # 4) Construir filas a partir de time entries
    excel_data = []
    task_ids_with_entries = set()
    for entry in (entries_response.data or []):
        task = entry.get("tasks", {})
        user_id = entry.get("user_id")
        user_info = user_dict.get(user_id, {}) if user_id else {}
        c_id = task.get("client_id")
        client_name = client_dict.get(c_id, "")
        rate = user_info.get("rate", 0) or 0
        duration = entry.get("duration", 0) or 0
        total = rate * duration
        task_ids_with_entries.add(task.get("id"))

        excel_data.append({
            "Abogado": user_info.get("username", ""),
            "Rol": user_info.get("role", ""),
            "Nombre del Cliente": client_name,
            "Asunto": task.get("title", ""),
            "Descripción": entry.get("description", ""),
            "Área": task.get("area", ""),
            "Tipo de facturación": "Por hora",
            "Tiempo reportado": format_hours_to_hhmm(duration),
            "Fecha de reporte": entry.get("start_time", "")[:10] if entry.get("start_time") else "",
            "Tarifa del abogado": format_currency(rate),
            "Total Tarifa x Tiempo": format_currency(total),
            "Estado de facturación": entry.get("facturado", ""),
            "Abogado asignado": task.get("assigned_user_name", "") or "Sin abogado asignado"
        })

    # 5) Agregar filas de tareas hourly sin tiempos
    for t in tasks_response.data:
        if t["id"] in task_ids_with_entries:
            continue
        client_name = client_dict.get(t.get("client_id"), "")
        excel_data.append({
            "Abogado": "",
            "Rol": "",
            "Nombre del Cliente": client_name,
            "Asunto": t.get("title", ""),
            "Descripción": "",
            "Área": t.get("area", ""),
            "Tipo de facturación": "Por hora",
            "Tiempo reportado": format_hours_to_hhmm(0),
            "Fecha de reporte": "",
            "Tarifa del abogado": format_currency(0),
            "Total Tarifa x Tiempo": format_currency(0),
            "Estado de facturación": "",
            "Abogado asignado": t.get("assigned_user_name", "") or "Sin abogado asignado"
        })

    return _create_comprehensive_excel_file(excel_data, "Reporte Hourly", start_date, end_date)


async def _generate_task_specific_report(start_date: datetime, end_date: datetime, task_id: int, report_type: str):
    """Generate a task-specific report based on the billing type"""
    
    # Get task information including billing type
    task_response = supabase.table("tasks").select("""
        id, title, client_id, area, billing_type, facturado, assigned_user_name,
        asesoria_tarif, total_value, monthly_limit_hours_tasks, permanent
    """).eq("id", task_id).single().execute()
    
    if not task_response.data:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    
    task = task_response.data
    billing_type = task.get("billing_type")
    
    # Get client information
    client_response = supabase.table("clients").select("id, name").eq("id", task["client_id"]).single().execute()
    if not client_response.data:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    client = client_response.data
    
    # Get time entries for the task in the date range
    time_entries_response = supabase.table("time_entries").select("""
        id, duration, start_time, end_time, description, user_id, facturado
    """).eq("task_id", task_id).gte("start_time", start_date.isoformat()).lte("end_time", end_date.isoformat()).execute()
    
    if not time_entries_response.data:
        raise HTTPException(status_code=404, detail="No hay registros de tiempo para esta tarea en el período especificado")
    
    # Get user information
    user_ids = list(set([entry["user_id"] for entry in time_entries_response.data]))
    users_response = supabase.table("users").select("id, username, role, cost_per_hour_client").in_("id", user_ids).execute()
    user_dict = {user["id"]: user for user in users_response.data}
    
    # Generate report based on billing type
    if billing_type == "hourly":
        return await _generate_task_hourly_report(task, client, time_entries_response.data, user_dict, start_date, end_date)
    elif billing_type == "tarifa_fija":
        return await _generate_task_fixed_rate_report(task, client, time_entries_response.data, user_dict, start_date, end_date)
    elif billing_type == "fija":  # mensualidad
        return await _generate_task_monthly_report(task, client, time_entries_response.data, user_dict, start_date, end_date)
    else:
        # Default to general format for unknown billing types
        return await _generate_task_general_report(task, client, time_entries_response.data, user_dict, start_date, end_date)


async def _generate_task_hourly_report(task: dict, client: dict, time_entries: List[dict], user_dict: dict, start_date: datetime, end_date: datetime):
    """Generate hourly billing report for a specific task"""
    
    excel_data = []
    total_hours = 0
    total_value = 0
    
    for entry in time_entries:
        user = user_dict.get(entry["user_id"], {})
        duration = entry.get("duration", 0) or 0
        rate = user.get("cost_per_hour_client", 0) or 0
        total = duration * rate
        
        total_hours += duration
        total_value += total
        
        excel_data.append({
            "Abogado": user.get("username", ""),
            "Rol": user.get("role", ""),
            "Cliente": client["name"],
            "Asunto": task["title"],
            "Trabajo": entry.get("description", ""),
            "Área": task.get("area", ""),
            "Fecha Trabajo": entry["start_time"][:10] if entry.get("start_time") else "",
            "Modo de Facturación": "Por Hora",
            "Tiempo Trabajado": format_hours_to_hhmm(duration),
            "Tarifa Horaria": format_currency(rate),
            "Moneda": "COP",
            "Total": format_currency(total),
            "Estado de facturación": entry.get("facturado", "no hay datos")
        })
    
    # Add summary row
    excel_data.append({
        "Abogado": "TOTAL",
        "Rol": "",
        "Cliente": "",
        "Asunto": "",
        "Trabajo": "",
        "Área": "",
        "Fecha Trabajo": "",
        "Modo de Facturación": "",
        "Tiempo Trabajado": format_hours_to_hhmm(total_hours),
        "Tarifa Horaria": "",
        "Moneda": "",
        "Total": format_currency(total_value),
        "Estado de facturación": ""
    })
    
    return _create_comprehensive_excel_file(excel_data, f"Reporte Tarea {task['id']} Por Hora", start_date, end_date)


async def _generate_task_fixed_rate_report(task: dict, client: dict, time_entries: List[dict], user_dict: dict, start_date: datetime, end_date: datetime):
    """Generate fixed rate billing report for a specific task"""
    
    excel_data = []
    total_hours = 0
    
    for entry in time_entries:
        user = user_dict.get(entry["user_id"], {})
        duration = entry.get("duration", 0) or 0
        total_hours += duration
        
        excel_data.append({
            "Abogado": user.get("username", ""),
            "Rol": user.get("role", ""),
            "Cliente": client["name"],
            "Asunto": task["title"],
            "Trabajo": entry.get("description", ""),
            "Área": task.get("area", ""),
            "Fecha Trabajo": entry["start_time"][:10] if entry.get("start_time") else "",
            "Modo de Facturación": "Tarifa Fija",
            "Tiempo Trabajado": format_hours_to_hhmm(duration),
            "Tarifa Fija": format_currency(task.get('total_value', 0)),
            "Moneda": "COP",
            "Total": format_currency(task.get('total_value', 0)),
            "Estado de facturación": entry.get("facturado", "no hay datos")
        })
    
    # Add summary row
    excel_data.append({
        "Abogado": "TOTAL",
        "Rol": "",
        "Cliente": "",
        "Asunto": "",
        "Trabajo": "",
        "Área": "",
        "Fecha Trabajo": "",
        "Modo de Facturación": "",
        "Tiempo Trabajado": format_hours_to_hhmm(total_hours),
        "Tarifa Fija": "",
        "Moneda": "",
        "Total": format_currency(task.get('total_value', 0)),
        "Estado de facturación": ""
    })
    
    return _create_comprehensive_excel_file(excel_data, f"Reporte Tarea {task['id']} Tarifa Fija", start_date, end_date)


async def _generate_task_monthly_report(task: dict, client: dict, time_entries: List[dict], user_dict: dict, start_date: datetime, end_date: datetime):
    """Generate monthly subscription billing report for a specific task"""
    
    excel_data = []
    total_hours = 0
    monthly_limit = task.get("monthly_limit_hours_tasks", 0)
    monthly_rate = task.get("asesoria_tarif", 0)
    
    for entry in time_entries:
        user = user_dict.get(entry["user_id"], {})
        duration = entry.get("duration", 0) or 0
        total_hours += duration
        
        excel_data.append({
            "Abogado": user.get("username", ""),
            "Rol": user.get("role", ""),
            "Cliente": client["name"],
            "Asunto": task["title"],
            "Trabajo": entry.get("description", ""),
            "Área": task.get("area", ""),
            "Fecha Trabajo": entry["start_time"][:10] if entry.get("start_time") else "",
            "Modo de Facturación": "Mensualidad",
            "Tiempo Trabajado": format_hours_to_hhmm(duration),
            "Límite Mensual": format_hours_to_hhmm(monthly_limit),
            "Tarifa Mensual": format_currency(monthly_rate),
            "Moneda": "COP",
            "Total": format_currency(monthly_rate),
            "Estado de facturación": entry.get("facturado", "no hay datos")
        })
    
    # Calculate additional charges if monthly limit is exceeded
    additional_charge = 0
    if total_hours > monthly_limit:
        excess_hours = total_hours - monthly_limit
        # Calculate additional charge based on user rates
        for entry in time_entries:
            user = user_dict.get(entry["user_id"], {})
            rate = user.get("cost_per_hour_client", 0) or 0
            duration = entry.get("duration", 0) or 0
            if total_hours > monthly_limit:
                if total_hours - duration >= monthly_limit:
                    # This entry contributes to additional charges
                    additional_charge += duration * rate
                else:
                    # Partial contribution to additional charges
                    partial_hours = total_hours - monthly_limit
                    additional_charge += partial_hours * rate
                total_hours -= duration
    
    # Add summary row
    excel_data.append({
        "Abogado": "TOTAL",
        "Rol": "",
        "Cliente": "",
        "Asunto": "",
        "Trabajo": "",
        "Área": "",
        "Fecha Trabajo": "",
        "Modo de Facturación": "",
        "Tiempo Trabajado": format_hours_to_hhmm(total_hours),
        "Límite Mensual": "",
        "Tarifa Mensual": "",
        "Moneda": "",
        "Total": format_currency(monthly_rate + additional_charge),
        "Estado de facturación": "",
        "Abogado asignado": ""
    })
    
    return _create_comprehensive_excel_file(excel_data, f"Reporte Tarea {task['id']} Mensualidad", start_date, end_date)


async def _generate_task_general_report(task: dict, client: dict, time_entries: List[dict], user_dict: dict, start_date: datetime, end_date: datetime):
    """Generate general format report for a specific task"""
    
    excel_data = []
    total_hours = 0
    total_value = 0
    
    for entry in time_entries:
        user = user_dict.get(entry["user_id"], {})
        duration = entry.get("duration", 0) or 0
        rate = user.get("cost_per_hour_client", 0) or 0
        total = duration * rate
        
        total_hours += duration
        total_value += total
        
        excel_data.append({
            "Abogado": user.get("username", ""),
            "Rol": user.get("role", ""),
            "Cliente": client["name"],
            "Asunto": task["title"],
            "Descripción": entry.get("description", ""),
            "Área": task.get("area", ""),
            "Tipo de facturación": get_billing_type_display(task.get("billing_type", "")),
            "Tiempo reportado": format_hours_to_hhmm(duration),
            "Fecha de reporte": entry["start_time"][:10] if entry.get("start_time") else "",
            "Tarifa del abogado": format_currency(rate),
            "Total Tarifa x Tiempo": format_currency(total),
            "Estado de facturación": entry.get("facturado", ""),
            "Abogado asignado": task.get("assigned_user_name", "") or "Sin abogado asignado"
        })
    
    # Add summary row
    excel_data.append({
        "Abogado": "TOTAL",
        "Rol": "",
        "Cliente": "",
        "Asunto": "",
        "Descripción": "",
        "Área": "",
        "Tipo de facturación": "",
        "Tiempo reportado": format_hours_to_hhmm(total_hours),
        "Fecha de reporte": "",
        "Tarifa del abogado": "",
        "Total Tarifa x Tiempo": format_currency(total_value),
        "Estado de facturación": "",
        "Abogado asignado": ""
    })
    
    return _create_comprehensive_excel_file(excel_data, f"Reporte Tarea {task['id']} General", start_date, end_date)


def _create_comprehensive_excel_file(data: List[dict], sheet_name: str, start_date: datetime, end_date: datetime, 
                                   conditional_formatting: bool = False, difference_column: str = None):
    """Create and return an Excel file with the comprehensive report data"""
    
    if not data:
        raise HTTPException(status_code=404, detail="No hay datos para generar el reporte")
    
    # Create DataFrame
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'fg_color': '#52b5f7',
            'border': 1
        })
        
        # Cell format
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        # Time format for duration columns
        time_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'num_format': '[hh]:mm'
        })
        
        # Conditional formatting for difference column (green/red text)
        if conditional_formatting and difference_column:
            green_format = workbook.add_format({'font_color': 'green', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            red_format = workbook.add_format({'font_color': 'red', 'border': 1, 'align': 'center', 'valign': 'vcenter'})
            
            # Apply conditional formatting to difference column
            diff_col_idx = ord(difference_column) - ord('A')
            worksheet.conditional_format(f'{difference_column}2:{difference_column}{len(df)+1}', {
                'type': 'cell',
                'criteria': '>',
                'value': 0,
                'format': green_format
            })
            worksheet.conditional_format(f'{difference_column}2:{difference_column}{len(df)+1}', {
                'type': 'cell',
                'criteria': '<',
                'value': 0,
                'format': red_format
            })
        
        # Apply headers and adjust columns
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(1, col_num, value, header_format)
            # Adjust column width based on content
            max_length = max(len(str(value)), df[value].astype(str).str.len().max())
            worksheet.set_column(col_num, col_num, min(max_length + 2, 30))
        
        # Title format
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Merge title row
        worksheet.merge_range(f'A1:{chr(ord("A") + len(df.columns) - 1)}1', f'{sheet_name}', title_format)
        
        # Apply cell formatting
        for row in range(len(df)):
            for col in range(len(df.columns)):
                value = df.iloc[row, col]
                col_name = df.columns[col]
                # Use conditional formatting if applicable, otherwise use regular cell format
                if conditional_formatting and difference_column and col == diff_col_idx:
                    # Skip as conditional formatting is already applied
                    pass
                else:
                    # Convert duration-like columns to Excel time with [hh]:mm format
                    if col_name in ("Tiempo reportado", "Tiempo Trabajado", "Límite de horas mensuales", "Límite Mensual"):
                        excel_time_value = None
                        if isinstance(value, (int, float)):
                            excel_time_value = float(value) / 24.0
                        elif isinstance(value, str) and ":" in value:
                            try:
                                parts = value.split(":")
                                hours = int(parts[0])
                                minutes = int(parts[1])
                                excel_time_value = (hours + minutes / 60.0) / 24.0
                            except Exception:
                                excel_time_value = None
                        if excel_time_value is not None:
                            worksheet.write_number(row + 2, col, excel_time_value, time_format)
                        else:
                            worksheet.write(row + 2, col, value, cell_format)
                    else:
                        worksheet.write(row + 2, col, value, cell_format)
    
    output.seek(0)
    
    # Generate filename
    filename = f"{sheet_name}_{start_date.date()}_{end_date.date()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/simplified_report")
async def generate_simplified_report(
    request: SimplifiedReportRequest,
    user: dict = Depends(role_required(["socio", "senior", "consultor"]))
):
    """
    Generate a simplified report showing all clients with tasks, including those with 0 time reported.
    Shows only the 7 required columns:
    1. Abogado asignado (assigned lawyer)
    2. Rol (role) 
    3. Nombre del cliente (client name)
    4. Asunto (subject/task title)
    5. Área (area)
    6. Tipo de facturación (billing type)
    7. Valor de facturación (billing value)
    """
    try:
        from datetime import datetime, date
        import calendar
        
        # Set default dates to current month if not provided
        if not request.start_date or not request.end_date:
            today = date.today()
            first_day = date(today.year, today.month, 1)
            last_day = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
            start_date = datetime.combine(first_day, datetime.min.time())
            end_date = datetime.combine(last_day, datetime.max.time())
        else:
            start_date = request.start_date
            end_date = request.end_date

        # Get all tasks for clients (regardless of time entries)
        query = supabase.table("tasks").select("""
            id, title, client_id, area, billing_type, assigned_user_name,
            asesoria_tarif, total_value, monthly_limit_hours_tasks
        """)
        
        # Only filter by client if client_id is provided
        if request.client_id:
            query = query.eq("client_id", request.client_id)
        
        tasks_response = query.execute()
        
        if not tasks_response.data:
            raise HTTPException(status_code=404, detail="No se encontraron tareas")
        
        # Get client information
        client_ids = list(set([task.get("client_id") for task in tasks_response.data if task.get("client_id")]))
        clients_response = supabase.table("clients").select("id, name").in_("id", client_ids).execute()
        client_dict = {client["id"]: client["name"] for client in clients_response.data} if clients_response.data else {}
        
        # Get user information for assigned lawyers
        assigned_users = list(set([task.get("assigned_user_name") for task in tasks_response.data if task.get("assigned_user_name")]))
        users_response = supabase.table("users").select("username, role").in_("username", assigned_users).execute()
        user_dict = {user["username"]: user["role"] for user in users_response.data} if users_response.data else {}
        
        # Prepare Excel data
        excel_data = []
        for task in tasks_response.data:
            client_name = client_dict.get(task.get("client_id"), "")
            assigned_lawyer = task.get("assigned_user_name", "") or "Sin abogado asignado"
            role = user_dict.get(assigned_lawyer, "") if assigned_lawyer != "Sin abogado asignado" else ""
            
            # Calculate billing value based on billing type
            billing_type = task.get("billing_type", "")
            billing_value = 0
            
            if billing_type == "tarifa_fija":
                billing_value = task.get("total_value", 0) or 0
            elif billing_type == "fija":  # mensualidad
                billing_value = task.get("asesoria_tarif", 0) or 0
            elif billing_type == "hourly":
                # For hourly tasks, we need to calculate based on time entries
                # Get time entries for this task in the date range
                time_entries_response = supabase.table("time_entries").select("""
                    duration, user_id
                """).eq("task_id", task["id"]).gte("start_time", start_date.isoformat()).lte("end_time", end_date.isoformat()).execute()
                
                if time_entries_response.data:
                    # Get user rates
                    user_ids = list(set([entry.get("user_id") for entry in time_entries_response.data if entry.get("user_id")]))
                    if user_ids:
                        users_rates_response = supabase.table("users").select("id, cost_per_hour_client").in_("id", user_ids).execute()
                        user_rates_dict = {user["id"]: user.get("cost_per_hour_client", 0) for user in users_rates_response.data} if users_rates_response.data else {}
                        
                        # Calculate total billing value
                        for entry in time_entries_response.data:
                            duration = entry.get("duration", 0) or 0
                            user_id = entry.get("user_id")
                            rate = user_rates_dict.get(user_id, 0) or 0
                            billing_value += duration * rate
                
                billing_value = round(billing_value, 2)
            
            excel_data.append({
                "Abogado asignado": assigned_lawyer,
                "Rol": role,
                "Nombre del cliente": client_name,
                "Asunto": task.get("title", ""),
                "Área": task.get("area", ""),
                "Tipo de facturación": get_billing_type_display(billing_type),
                "Valor de facturación": format_currency(billing_value)
            })
        
        return _create_simplified_excel_file(excel_data, "Reporte Clientes Totales", start_date, end_date)

    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")


def _create_simplified_excel_file(data: List[dict], sheet_name: str, start_date: datetime, end_date: datetime):
    """Create and return an Excel file with the simplified report data"""
    
    if not data:
        raise HTTPException(status_code=404, detail="No hay datos para generar el reporte")
    
    # Create DataFrame
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'vcenter',
            'align': 'center',
            'fg_color': '#D7E4BC',
            'border': 1
        })
        
        # Cell format
        cell_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Apply headers and adjust columns
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(1, col_num, value, header_format)
            # Adjust column width based on content
            max_length = max(len(str(value)), df[value].astype(str).str.len().max())
            worksheet.set_column(col_num, col_num, min(max_length + 2, 30))
        
        # Title format
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Merge title row
        worksheet.merge_range(f'A1:{chr(ord("A") + len(df.columns) - 1)}1', f'{sheet_name}', title_format)
        
        # Apply cell formatting
        for row in range(len(df)):
            for col in range(len(df.columns)):
                value = df.iloc[row, col]
                worksheet.write(row + 2, col, value, cell_format)
    
    output.seek(0)
    
    # Generate filename
    filename = f"{sheet_name}_{start_date.date()}_{end_date.date()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
